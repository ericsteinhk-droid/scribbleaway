"""
File Search GUI — searches for keywords across Word (.docx) OR Excel (.xlsx/.xls) files.
User selects the file type at startup.
Requires: python-docx, openpyxl, xlrd, Pillow
Build to .exe: pyinstaller file_search.spec
"""

import csv
import os
import queue
import re
import sys
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# ---- optional dependencies (checked at startup) ----
try:
    from docx import Document as _DocxDoc
    from docx.opc.exceptions import PackageNotFoundError as _DocxErr
    _DOCX_OK = True
except ImportError:
    _DOCX_OK = False

try:
    import openpyxl as _openpyxl
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

try:
    import xlrd as _xlrd
    _XLRD_OK = True
except ImportError:
    _XLRD_OK = False

try:
    from PIL import Image, ImageTk
    _PIL_OK = True
except ImportError:
    _PIL_OK = False


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
APP_VERSION  = "v5.1"
APP_DATE     = "May 2026"
COPYRIGHT    = f"© Eric Stein, EVOQ Architecture  ·  {APP_VERSION}  ·  {APP_DATE}"

_COL_OPEN  = 26
_COL_PATH  = 36
_COL_KW    = 18
_COL_EXCPT = 48

FILE_TYPES = {
    "word":  {"label": "Word Documents",        "exts": (".docx",),          "color": "#1a5276"},
    "excel": {"label": "Excel Spreadsheets",    "exts": (".xlsx", ".xlsm", ".xls"), "color": "#1e8449"},
}


# ---------------------------------------------------------------------------
# Resource path (dev vs PyInstaller bundle)
# ---------------------------------------------------------------------------
def _resource_path(relative: str) -> str:
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, relative)


# ---------------------------------------------------------------------------
# Text extraction
# ---------------------------------------------------------------------------

def _extract_docx(path: str) -> str:
    doc = _DocxDoc(path)
    parts = [p.text for p in doc.paragraphs]
    for table in doc.tables:
        for row in table.rows:
            for cell in row.cells:
                parts.append(cell.text)
    return "\n".join(parts)


def _extract_xlsx(path: str) -> str:
    wb = _openpyxl.load_workbook(path, read_only=True, data_only=True)
    parts = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    parts.append(str(cell))
    wb.close()
    return "\n".join(parts)


def _extract_xls(path: str) -> str:
    wb = _xlrd.open_workbook(path)
    parts = []
    for sheet in wb.sheets():
        for rx in range(sheet.nrows):
            for cx in range(sheet.ncols):
                v = sheet.cell(rx, cx).value
                if v:
                    parts.append(str(v))
    return "\n".join(parts)


def _extract_text(path: str) -> str:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".docx":
        return _extract_docx(path)
    if ext in (".xlsx", ".xlsm"):
        return _extract_xlsx(path)
    if ext == ".xls":
        return _extract_xls(path)
    raise ValueError(f"Unsupported extension: {ext}")


# ---------------------------------------------------------------------------
# File collection + excerpt helper
# ---------------------------------------------------------------------------

def _collect_files(folder: str, recursive: bool, exts: tuple):
    if recursive:
        for root, _dirs, files in os.walk(folder):
            for f in files:
                if os.path.splitext(f)[1].lower() in exts:
                    yield os.path.join(root, f)
    else:
        for f in os.listdir(folder):
            if os.path.splitext(f)[1].lower() in exts:
                yield os.path.join(folder, f)


def _sentences_around(text: str, keyword: str, case_sensitive: bool, context: int = 90) -> str:
    flags = 0 if case_sensitive else re.IGNORECASE
    match = re.search(re.escape(keyword), text, flags=flags)
    if not match:
        return ""
    kw_start, kw_end = match.start(), match.end()
    start = max(0, kw_start - context)
    end   = min(len(text), kw_end + context)
    if start == 0 and kw_start < context:
        end = min(len(text), end + (context - kw_start))
    if end == len(text) and (len(text) - kw_end) < context:
        start = max(0, start - (context - (len(text) - kw_end)))
    excerpt = text[start:end].replace("\n", " ").strip()
    if start > 0:
        excerpt = "…" + excerpt
    if end < len(text):
        excerpt = excerpt + "…"
    return excerpt


# ---------------------------------------------------------------------------
# Search worker
# ---------------------------------------------------------------------------

def search_worker(folder, recursive, exts, keywords, match_all, case_sensitive, result_queue):
    paths = list(_collect_files(folder, recursive, exts))
    total = len(paths)
    match_count = 0
    files_with_matches = set()

    for idx, path in enumerate(paths, 1):
        filename = os.path.basename(path)
        result_queue.put({"type": "progress", "current": idx, "total": total, "filename": filename})

        try:
            text = _extract_text(path)
        except Exception as exc:
            result_queue.put({"type": "error", "filename": filename, "path": path, "reason": str(exc)})
            continue

        flags = 0 if case_sensitive else re.IGNORECASE
        matched = [kw for kw in keywords if re.search(re.escape(kw), text, flags)]
        hit = (len(matched) == len(keywords)) if match_all else bool(matched)

        if hit:
            for kw in matched:
                result_queue.put({
                    "type": "match",
                    "filename": filename,
                    "path": path,
                    "keyword": kw,
                    "excerpt": _sentences_around(text, kw, case_sensitive),
                })
                match_count += 1
            files_with_matches.add(path)

    result_queue.put({"type": "done", "matches": match_count, "files_with_matches": files_with_matches})


# ---------------------------------------------------------------------------
# Scrollable frame
# ---------------------------------------------------------------------------

class _ScrollableFrame(tk.Frame):
    def __init__(self, master, **kw):
        super().__init__(master, **kw)
        self._canvas = tk.Canvas(self, bg="#ffffff", highlightthickness=0)
        vsb = ttk.Scrollbar(self, orient="vertical",   command=self._canvas.yview)
        hsb = ttk.Scrollbar(self, orient="horizontal", command=self._canvas.xview)
        self._canvas.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._canvas.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)
        self.inner = tk.Frame(self._canvas, bg="#ffffff")
        self._win  = self._canvas.create_window((0, 0), window=self.inner, anchor="nw")
        self.inner.bind("<Configure>", lambda _e: self._canvas.configure(
            scrollregion=self._canvas.bbox("all")))
        self._canvas.bind("<Configure>", lambda e: self._canvas.itemconfig(
            self._win, width=e.width))
        self._canvas.bind("<Enter>", lambda _e: self._canvas.bind_all(
            "<MouseWheel>", lambda e: self._canvas.yview_scroll(int(-1*(e.delta/120)), "units")))
        self._canvas.bind("<Leave>", lambda _e: self._canvas.unbind_all("<MouseWheel>"))

    def clear(self):
        for w in self.inner.winfo_children():
            w.destroy()


# ---------------------------------------------------------------------------
# Startup: file-type chooser
# ---------------------------------------------------------------------------

class FileTypeChooser(tk.Tk):
    """Small launcher window — user picks Word or Excel before the main app opens."""

    def __init__(self):
        super().__init__()
        self.result = None
        self.title("EVOQ File Search")
        self.resizable(False, False)
        self.configure(bg="#ffffff")
        self._build()
        self.eval("tk::PlaceWindow . center")

    def _build(self):
        # Logo
        if _PIL_OK:
            try:
                img = Image.open(_resource_path("evoq_logo.png")).convert("RGBA")
                h = 48
                img = img.resize((int(img.width * h / img.height), h), Image.LANCZOS)
                self._logo = ImageTk.PhotoImage(img)
                tk.Label(self, image=self._logo, bg="#ffffff").pack(pady=(18, 4))
            except Exception:
                pass

        tk.Label(self, text="File Search", font=("Segoe UI", 16, "bold"),
                 fg="#1a237e", bg="#ffffff").pack()
        tk.Label(self, text="What type of files do you want to search?",
                 font=("Segoe UI", 10), fg="#555", bg="#ffffff").pack(pady=(4, 20))

        btn_frame = tk.Frame(self, bg="#ffffff")
        btn_frame.pack(padx=30, pady=(0, 10))

        self._make_tile(btn_frame, "Word Documents", ".docx", "#1a5276", "#d6eaf8",
                        "word").grid(row=0, column=0, padx=10, pady=6)
        self._make_tile(btn_frame, "Excel Spreadsheets", ".xlsx / .xls", "#1e8449", "#d5f5e3",
                        "excel").grid(row=0, column=1, padx=10, pady=6)

        tk.Label(self, text=COPYRIGHT, font=("Segoe UI", 7),
                 fg="#c0c0c0", bg="#ffffff").pack(pady=(8, 10))

    def _make_tile(self, parent, title, subtitle, fg, bg, file_type):
        tile = tk.Frame(parent, bg=bg, cursor="hand2", relief="flat", bd=0,
                        width=160, height=110)
        tile.pack_propagate(False)
        tk.Label(tile, text=title, font=("Segoe UI", 11, "bold"),
                 fg=fg, bg=bg, wraplength=140).pack(pady=(22, 2))
        tk.Label(tile, text=subtitle, font=("Segoe UI", 9),
                 fg="#555", bg=bg).pack()
        for widget in (tile, *tile.winfo_children()):
            widget.bind("<Button-1>", lambda _e, ft=file_type: self._choose(ft))
            widget.bind("<Enter>",    lambda _e, t=tile, c=fg:  t.config(bg=c) or
                        [w.config(bg=c, fg="#ffffff") for w in t.winfo_children()])
            widget.bind("<Leave>",   lambda _e, t=tile, c=bg, f=fg: t.config(bg=c) or
                        [w.config(bg=c, fg=f if w.cget("font").find("bold") >= 0 else "#555")
                         for w in t.winfo_children()])
        return tile

    def _choose(self, file_type):
        self.result = file_type
        self.destroy()


# ---------------------------------------------------------------------------
# Main search window
# ---------------------------------------------------------------------------

class SearchApp(tk.Tk):
    def __init__(self, file_type: str):
        super().__init__()
        self._ft    = file_type
        self._meta  = FILE_TYPES[file_type]
        self._exts  = self._meta["exts"]
        self._color = self._meta["color"]
        self._label = self._meta["label"]

        self.title(f"EVOQ File Search — {self._label}")
        self.minsize(960, 620)
        self.configure(bg="#f0f0f0")

        self._result_queue  = queue.Queue()
        self._search_thread = None
        self._all_rows      = []
        self._error_rows    = []
        self._row_count     = 0

        self._build_ui()
        self.after(100, self._poll_queue)

    # ------------------------------------------------------------------
    # UI
    # ------------------------------------------------------------------

    def _build_ui(self):
        pad = {"padx": 8, "pady": 4}

        # ---- Logo header ----
        hbar = tk.Frame(self, bg="#ffffff", height=64)
        hbar.pack(fill="x")
        hbar.pack_propagate(False)
        if _PIL_OK:
            try:
                img = Image.open(_resource_path("evoq_logo.png")).convert("RGBA")
                h = 46
                img = img.resize((int(img.width * h / img.height), h), Image.LANCZOS)
                self._logo = ImageTk.PhotoImage(img)
                tk.Label(hbar, image=self._logo, bg="#ffffff").pack(side="left", padx=14, pady=9)
            except Exception:
                pass

        tk.Label(hbar, text="File Search", font=("Segoe UI", 15, "bold"),
                 fg="#1a237e", bg="#ffffff").pack(side="left", padx=(4, 0))
        tk.Label(hbar, text=f"— {self._label}",
                 font=("Segoe UI", 10), fg=self._color, bg="#ffffff").pack(
                 side="left", padx=(6, 0), pady=(6, 0))
        tk.Frame(self, bg="#e0e0e0", height=1).pack(fill="x")

        # ---- Settings ----
        top = tk.LabelFrame(self, text="Search Settings", bg="#f0f0f0",
                            font=("Segoe UI", 9, "bold"))
        top.pack(fill="x", padx=10, pady=(10, 4))

        tk.Label(top, text="Folder:", bg="#f0f0f0").grid(row=0, column=0, sticky="e", **pad)
        self._folder_var = tk.StringVar()
        tk.Entry(top, textvariable=self._folder_var, width=60).grid(row=0, column=1, sticky="ew", **pad)
        tk.Button(top, text="Browse…", command=self._browse_folder).grid(row=0, column=2, **pad)

        tk.Label(top, text="Keywords:", bg="#f0f0f0").grid(row=1, column=0, sticky="e", **pad)
        self._keywords_var = tk.StringVar()
        tk.Entry(top, textvariable=self._keywords_var, width=60).grid(row=1, column=1, sticky="ew", **pad)
        tk.Label(top, text="(comma-separated)", bg="#f0f0f0", fg="#666").grid(
            row=1, column=2, sticky="w", **pad)

        opts = tk.Frame(top, bg="#f0f0f0")
        opts.grid(row=2, column=0, columnspan=3, sticky="w", padx=8, pady=4)
        self._recursive_var = tk.BooleanVar(value=True)
        tk.Checkbutton(opts, text="Include subfolders", variable=self._recursive_var,
                       bg="#f0f0f0").pack(side="left", padx=(0, 16))
        self._match_all_var = tk.BooleanVar(value=False)
        tk.Checkbutton(opts, text="Match ALL keywords (AND)", variable=self._match_all_var,
                       bg="#f0f0f0").pack(side="left", padx=(0, 16))
        self._case_var = tk.BooleanVar(value=False)
        tk.Checkbutton(opts, text="Case-sensitive", variable=self._case_var,
                       bg="#f0f0f0").pack(side="left")
        top.columnconfigure(1, weight=1)

        # ---- Action bar ----
        action = tk.Frame(self, bg="#f0f0f0")
        action.pack(fill="x", padx=10, pady=2)
        self._search_btn = tk.Button(
            action, text="Search", bg=self._color, fg="white",
            font=("Segoe UI", 10, "bold"), padx=16, pady=4,
            relief="flat", cursor="hand2", command=self._start_search)
        self._search_btn.pack(side="left")
        self._export_btn = tk.Button(action, text="Export to CSV…", padx=12, pady=4,
            relief="flat", cursor="hand2", command=self._export_csv, state="disabled")
        self._export_btn.pack(side="left", padx=(8, 0))
        self._errors_btn = tk.Button(action, text="Show Skipped Files", padx=12, pady=4,
            relief="flat", cursor="hand2", command=self._show_errors, state="disabled")
        self._errors_btn.pack(side="left", padx=(8, 0))
        self._status_var = tk.StringVar(value="Ready.")
        tk.Label(action, textvariable=self._status_var, bg="#f0f0f0", fg="#444",
                 font=("Segoe UI", 9)).pack(side="left", padx=16)

        # ---- Progress ----
        self._progress_var = tk.DoubleVar(value=0)
        ttk.Progressbar(self, variable=self._progress_var, maximum=100).pack(
            fill="x", padx=10, pady=(2, 4))

        # ---- Footer ----
        tk.Label(self, text=COPYRIGHT, font=("Segoe UI", 7),
                 fg="#b0b0b0", bg="#f0f0f0").pack(side="bottom", pady=(0, 3))

        # ---- Results ----
        rf = tk.LabelFrame(self, text="Results", bg="#f0f0f0", font=("Segoe UI", 9, "bold"))
        rf.pack(fill="both", expand=True, padx=10, pady=(0, 4))
        rf.rowconfigure(1, weight=1)
        rf.columnconfigure(0, weight=1)

        hdr = tk.Frame(rf, bg="#c8d4e8")
        hdr.grid(row=0, column=0, sticky="ew")
        _hf = ("Segoe UI", 9, "bold")
        tk.Label(hdr, text="Open File", font=_hf, bg="#c8d4e8",
                 width=_COL_OPEN, anchor="w").pack(side="left", padx=(6, 2), pady=3)
        tk.Label(hdr, text="Full Path", font=_hf, bg="#c8d4e8",
                 width=_COL_PATH, anchor="w").pack(side="left", padx=2, pady=3)
        tk.Label(hdr, text="Matched Keyword", font=_hf, bg="#c8d4e8",
                 width=_COL_KW, anchor="w").pack(side="left", padx=2, pady=3)
        tk.Label(hdr, text="Excerpt", font=_hf, bg="#c8d4e8",
                 width=_COL_EXCPT, anchor="w").pack(side="left", padx=2, pady=3)

        self._results_sf = _ScrollableFrame(rf, bg="#ffffff")
        self._results_sf.grid(row=1, column=0, sticky="nsew")

    # ------------------------------------------------------------------
    # Row builder
    # ------------------------------------------------------------------

    def _add_result_row(self, filename, path, keyword, excerpt):
        self._row_count += 1
        bg = "#ffffff" if self._row_count % 2 == 1 else "#f0f4fa"
        row = tk.Frame(self._results_sf.inner, bg=bg)
        row.pack(fill="x")

        tk.Button(
            row, text=filename,
            font=("Segoe UI", 9, "bold"),
            bg="#FFE066", fg="#1a237e",
            activebackground="#FFD600", activeforeground="#1a237e",
            relief="flat", bd=0, cursor="hand2",
            anchor="w", width=_COL_OPEN, padx=4,
            command=lambda p=path: self._open_file(p),
        ).pack(side="left", padx=(4, 2), pady=2)

        _lf = ("Segoe UI", 9)
        tk.Label(row, text=path,    font=_lf, bg=bg, width=_COL_PATH, anchor="w",
                 wraplength=260).pack(side="left", padx=2, pady=2)
        tk.Label(row, text=keyword, font=_lf, bg=bg, width=_COL_KW,   anchor="w").pack(
                 side="left", padx=2, pady=2)
        self._make_excerpt_widget(row, excerpt, keyword, bg).pack(side="left", padx=2, pady=2)

    def _make_excerpt_widget(self, parent, excerpt, keyword, bg):
        w = tk.Text(parent, font=("Segoe UI", 9), bg=bg, relief="flat", bd=0,
                    highlightthickness=0, wrap="word", height=2,
                    width=_COL_EXCPT, cursor="arrow")
        w.tag_configure("bold", font=("Segoe UI", 9, "bold"))
        for part in re.split(f"({re.escape(keyword)})", excerpt, flags=re.IGNORECASE):
            if re.fullmatch(re.escape(keyword), part, flags=re.IGNORECASE):
                w.insert("end", part, "bold")
            else:
                w.insert("end", part)
        w.config(state="disabled")
        return w

    def _open_file(self, path):
        if os.path.exists(path):
            os.startfile(path)
        else:
            messagebox.showwarning("File not found", f"Cannot open:\n{path}")

    # ------------------------------------------------------------------
    # Search
    # ------------------------------------------------------------------

    def _browse_folder(self):
        folder = filedialog.askdirectory(title="Select folder to search")
        if folder:
            self._folder_var.set(folder)

    def _start_search(self):
        folder = self._folder_var.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showwarning("No folder", "Please select a valid folder first.")
            return
        keywords = [k.strip() for k in self._keywords_var.get().split(",") if k.strip()]
        if not keywords:
            messagebox.showwarning("No keywords", "Please enter at least one keyword.")
            return
        if self._search_thread and self._search_thread.is_alive():
            messagebox.showinfo("Busy", "A search is already running. Please wait.")
            return

        self._all_rows.clear()
        self._error_rows.clear()
        self._row_count = 0
        self._results_sf.clear()
        self._progress_var.set(0)
        self._export_btn.config(state="disabled")
        self._errors_btn.config(state="disabled")
        self._search_btn.config(state="disabled")
        self._status_var.set("Scanning…")

        self._search_thread = threading.Thread(
            target=search_worker,
            args=(folder, self._recursive_var.get(), self._exts,
                  keywords, self._match_all_var.get(), self._case_var.get(),
                  self._result_queue),
            daemon=True,
        )
        self._search_thread.start()

    def _poll_queue(self):
        try:
            while True:
                msg = self._result_queue.get_nowait()
                t   = msg["type"]
                if t == "progress":
                    pct = (msg["current"] / msg["total"] * 100) if msg["total"] else 0
                    self._progress_var.set(pct)
                    self._status_var.set(
                        f"Scanning {msg['current']}/{msg['total']}: {msg['filename']}")
                elif t == "match":
                    self._all_rows.append(
                        (msg["filename"], msg["path"], msg["keyword"], msg["excerpt"]))
                    self._add_result_row(
                        msg["filename"], msg["path"], msg["keyword"], msg["excerpt"])
                elif t == "error":
                    self._error_rows.append((msg["filename"], msg["path"], msg["reason"]))
                elif t == "done":
                    n, f = msg["matches"], len(msg["files_with_matches"])
                    self._status_var.set(
                        f"Done. Found {n} match{'es' if n != 1 else ''}"
                        f" across {f} file{'s' if f != 1 else ''}.")
                    self._progress_var.set(100)
                    self._search_btn.config(state="normal")
                    if self._all_rows:   self._export_btn.config(state="normal")
                    if self._error_rows: self._errors_btn.config(state="normal")
        except queue.Empty:
            pass
        self.after(50, self._poll_queue)

    # ------------------------------------------------------------------
    # Export / errors
    # ------------------------------------------------------------------

    def _export_csv(self):
        if not self._all_rows:
            messagebox.showinfo("Nothing to export", "No results to export.")
            return
        dest = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
        if not dest:
            return
        try:
            with open(dest, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(["File Name", "Full Path", "Matched Keyword", "Excerpt"])
                w.writerows(self._all_rows)
            messagebox.showinfo("Exported", f"Results saved to:\n{dest}")
        except OSError as exc:
            messagebox.showerror("Export failed", str(exc))

    def _show_errors(self):
        if not self._error_rows:
            return
        win = tk.Toplevel(self)
        win.title("Skipped / Unreadable Files")
        win.minsize(700, 300)
        cols = ("filename", "path", "reason")
        tree = ttk.Treeview(win, columns=cols, show="headings")
        tree.heading("filename", text="File Name")
        tree.heading("path",     text="Full Path")
        tree.heading("reason",   text="Reason Skipped")
        tree.column("filename", width=160)
        tree.column("path",     width=280)
        tree.column("reason",   width=240)
        vsb = ttk.Scrollbar(win, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        for row in self._error_rows:
            tree.insert("", "end", values=row)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    # Dependency check
    missing = []
    if not _DOCX_OK:
        missing.append("python-docx  (pip install python-docx)")
    if not _OPENPYXL_OK:
        missing.append("openpyxl     (pip install openpyxl)")
    if missing:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Missing dependencies",
                             "Please install:\n\n" + "\n".join(missing))
        raise SystemExit(1)

    chooser = FileTypeChooser()
    chooser.mainloop()

    if chooser.result:
        app = SearchApp(chooser.result)
        app.mainloop()
